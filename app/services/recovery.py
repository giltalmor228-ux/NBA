from __future__ import annotations

import base64
import csv
import io
import json
import tempfile
import zipfile
from collections.abc import Iterable
from decimal import Decimal
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.domain.scoring import MemberState, ResultEnvelope, SubmissionEnvelope, WindowEnvelope, leaderboard_as_dict, score_pool
from app.models import BettingWindow, EventLog, InviteLink, Membership, PaymentLedgerEntry, PickSubmission, Pool, ResultSnapshot, User


TABLES = [
    ("pools", Pool),
    ("users", User),
    ("memberships", Membership),
    ("invite_links", InviteLink),
    ("betting_windows", BettingWindow),
    ("pick_submissions", PickSubmission),
    ("result_snapshots", ResultSnapshot),
    ("event_logs", EventLog),
    ("payment_ledger_entries", PaymentLedgerEntry),
]

BYTES_PREFIX = "base64:"


def _to_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (bytes, bytearray)):
        return f"{BYTES_PREFIX}{base64.b64encode(bytes(value)).decode('ascii')}"
    return value


def _to_json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: _to_json_safe(val) for key, val in value.items()}
    if isinstance(value, list):
        return [_to_json_safe(item) for item in value]
    return _to_value(value)


def _row_dict(instance: Any) -> dict[str, Any]:
    return {column.name: _to_value(getattr(instance, column.name)) for column in instance.__table__.columns}


def _from_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(BYTES_PREFIX):
        return base64.b64decode(value.removeprefix(BYTES_PREFIX))
    return value


def build_snapshot(session: Session, pool_id: str) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    pool = session.get(Pool, pool_id)
    payload["exported_at"] = datetime.now(timezone.utc).isoformat()
    payload["pool_id"] = pool_id
    payload["pool"] = _row_dict(pool)
    for name, model in TABLES[1:]:
        stmt = select(model)
        if hasattr(model, "pool_id"):
            stmt = stmt.where(model.pool_id == pool_id)
        elif model is User:
            membership_user_ids = select(Membership.user_id).where(Membership.pool_id == pool_id)
            stmt = stmt.where(model.id.in_(membership_user_ids))
        payload[name] = [_row_dict(item) for item in session.scalars(stmt).all()]
    payload["leaderboard"] = _to_json_safe(leaderboard_as_dict(_load_leaderboard(session, pool_id)))
    return payload


def _load_leaderboard(session: Session, pool_id: str):
    memberships = session.scalars(select(Membership).where(Membership.pool_id == pool_id)).all()
    users = {user.id: user for user in session.scalars(select(User).where(User.id.in_([m.user_id for m in memberships]))).all()}
    windows = session.scalars(select(BettingWindow).where(BettingWindow.pool_id == pool_id)).all()
    submissions = session.scalars(
        select(PickSubmission).where(PickSubmission.window_id.in_([window.id for window in windows]))
    ).all()
    results = session.scalars(select(ResultSnapshot).where(ResultSnapshot.pool_id == pool_id)).all()
    return score_pool(
        members=[
            MemberState(
                member_id=membership.id,
                display_name=users[membership.user_id].nickname,
                payout_eligible=membership.payout_eligible,
                is_monkey=users[membership.user_id].is_monkey,
            )
            for membership in memberships
        ],
        windows=[
            WindowEnvelope(
                window_id=window.id,
                name=window.name,
                round_key=window.round_key,
                bet_type=window.bet_type,
                config=window.config,
                is_locked=window.is_locked,
            )
            for window in windows
        ],
        submissions=[
            SubmissionEnvelope(
                window_id=submission.window_id,
                member_id=submission.member_id,
                submitted_at=submission.submitted_at,
                payload=submission.payload,
            )
            for submission in submissions
        ],
        results=[
            ResultEnvelope(
                scope_type=result.scope_type,
                scope_key=result.scope_key,
                created_at=result.created_at,
                payload=result.payload,
                is_override=result.is_override,
            )
            for result in results
        ],
    )


def _write_csv(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    rows = list(rows)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _build_workbook(snapshot: dict[str, Any]) -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions["A1"] = "Authoritative recovery source"
    instructions["B1"] = "snapshot.json in the export bundle"
    instructions["A3"] = "Workbook role"
    instructions["B3"] = "Operator-friendly fallback view with formulas and current standings"

    leaderboard = workbook.create_sheet("Leaderboard")
    leaderboard.append(["Rank", "Payout Rank", "Player", "Points", "Exact Hits", "FMVP Correct", "Ceiling"])
    for row in snapshot["leaderboard"]:
        leaderboard.append(
            [
                row["rank"],
                row["payout_rank"],
                row["display_name"],
                row["total_points"],
                row["exact_hits"],
                "YES" if row["finals_mvp_correct"] else "NO",
                row["projected_ceiling"],
            ]
        )
    leaderboard["I1"] = "Visible totals"
    leaderboard["I2"] = "=SUM(D2:D200)"

    for sheet_name in ("memberships", "betting_windows", "pick_submissions", "result_snapshots", "payment_ledger_entries"):
        sheet = workbook.create_sheet(sheet_name[:31])
        rows = snapshot.get(sheet_name, [])
        if not rows:
            continue
        sheet.append(list(rows[0].keys()))
        for row in rows:
            sheet.append([json.dumps(value) if isinstance(value, (dict, list)) else value for value in row.values()])

    handle = io.BytesIO()
    workbook.save(handle)
    return handle.getvalue()


def export_bundle(session: Session, pool_id: str) -> tuple[str, bytes]:
    snapshot = build_snapshot(session, pool_id)
    filename = f"pool-backup-{pool_id}.zip"
    with tempfile.TemporaryDirectory() as temp_dir:
        root = Path(temp_dir)
        (root / "snapshot.json").write_text(json.dumps(snapshot, indent=2), encoding="utf-8")
        csv_dir = root / "csv"
        csv_dir.mkdir()
        for key, rows in snapshot.items():
            if isinstance(rows, list) and rows and isinstance(rows[0], dict):
                _write_csv(csv_dir / f"{key}.csv", rows)
        (root / "fallback_workbook.xlsx").write_bytes(_build_workbook(snapshot))
        archive_buffer = io.BytesIO()
        with zipfile.ZipFile(archive_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
            for file_path in root.rglob("*"):
                archive.write(file_path, file_path.relative_to(root))
    return filename, archive_buffer.getvalue()


def restore_from_snapshot_json(session: Session, raw_bytes: bytes) -> tuple[Pool, str]:
    snapshot = json.loads(raw_bytes.decode("utf-8"))
    pool = Pool(name=f"{snapshot['pool']['name']} (Recovered)", season_label=snapshot["pool"]["season_label"])
    session.add(pool)
    session.flush()

    user_map: dict[str, str] = {}
    member_map: dict[str, str] = {}
    commissioner_membership_id = ""

    for user_row in snapshot["users"]:
        user = User(
            email=user_row.get("email"),
            nickname=user_row["nickname"],
            avatar=user_row["avatar"],
            loser_photo_path=user_row.get("loser_photo_path"),
            loser_photo_content_type=user_row.get("loser_photo_content_type"),
            loser_photo_blob=_from_value(user_row.get("loser_photo_blob")),
            is_monkey=user_row["is_monkey"],
        )
        session.add(user)
        session.flush()
        user_map[user_row["id"]] = user.id

    for member_row in snapshot["memberships"]:
        membership = Membership(
            pool_id=pool.id,
            user_id=user_map[member_row["user_id"]],
            role=member_row["role"],
            payout_eligible=member_row["payout_eligible"],
            payment_status=member_row["payment_status"],
        )
        session.add(membership)
        session.flush()
        member_map[member_row["id"]] = membership.id
        if membership.role == "commissioner" and not commissioner_membership_id:
            commissioner_membership_id = membership.id

    for invite_row in snapshot["invite_links"]:
        session.add(InviteLink(pool_id=pool.id, token=f"recovered-{invite_row['token']}", active=invite_row["active"]))

    window_map: dict[str, str] = {}
    for window_row in snapshot["betting_windows"]:
        window = BettingWindow(
            pool_id=pool.id,
            name=window_row["name"],
            round_key=window_row["round_key"],
            bet_type=window_row["bet_type"],
            config=window_row["config"],
            opens_at=datetime.fromisoformat(window_row["opens_at"]),
            locks_at=datetime.fromisoformat(window_row["locks_at"]),
            is_locked=window_row["is_locked"],
            is_revealed=window_row["is_revealed"],
            revealed_at=datetime.fromisoformat(window_row["revealed_at"]) if window_row["revealed_at"] else None,
            monkey_seed=window_row["monkey_seed"],
        )
        session.add(window)
        session.flush()
        window_map[window_row["id"]] = window.id

    def remap_payload(value: Any) -> Any:
        if isinstance(value, dict):
            return {key: remap_payload(val) for key, val in value.items()}
        if isinstance(value, list):
            return [remap_payload(item) for item in value]
        if isinstance(value, str) and value in member_map:
            return member_map[value]
        return value

    for submission_row in snapshot["pick_submissions"]:
        session.add(
            PickSubmission(
                window_id=window_map[submission_row["window_id"]],
                member_id=member_map[submission_row["member_id"]],
                payload=remap_payload(submission_row["payload"]),
                submitted_at=datetime.fromisoformat(submission_row["submitted_at"]),
            )
        )

    for result_row in snapshot["result_snapshots"]:
        session.add(
            ResultSnapshot(
                pool_id=pool.id,
                scope_type=result_row["scope_type"],
                scope_key=result_row["scope_key"],
                payload=remap_payload(result_row["payload"]),
                source=result_row["source"],
                is_override=result_row["is_override"],
                override_reason=result_row["override_reason"],
                created_by_member_id=member_map.get(result_row["created_by_member_id"]),
                created_at=datetime.fromisoformat(result_row["created_at"]),
            )
        )

    for event_row in snapshot["event_logs"]:
        session.add(
            EventLog(
                pool_id=pool.id,
                actor_member_id=member_map.get(event_row["actor_member_id"]),
                event_type=event_row["event_type"],
                payload=remap_payload(event_row["payload"]),
                created_at=datetime.fromisoformat(event_row["created_at"]),
            )
        )

    for payment_row in snapshot["payment_ledger_entries"]:
        session.add(
            PaymentLedgerEntry(
                pool_id=pool.id,
                member_id=member_map[payment_row["member_id"]],
                status=payment_row["status"],
                amount=payment_row["amount"],
                note=payment_row["note"],
                created_at=datetime.fromisoformat(payment_row["created_at"]),
            )
        )

    session.commit()
    return pool, commissioner_membership_id


def restore_from_workbook(workbook_bytes: bytes) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    data: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in ("memberships", "betting_windows", "pick_submissions", "result_snapshots", "payment_ledger_entries"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = rows[0]
        data[sheet_name] = [dict(zip(headers, row, strict=False)) for row in rows[1:] if any(cell is not None for cell in row)]
    return data
